#!/usr/bin/env python3
"""Audit exact benchmark overlap against PRIME2 Table S4 without redistributing it."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import defaultdict
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def normalize_hla(value: object) -> str:
    compact = re.sub(r"[^A-Z0-9]", "", str(value).upper().removeprefix("HLA"))
    match = re.fullmatch(r"([ABC])(\d{2})(\d{2})", compact)
    if not match:
        raise ValueError(f"unsupported HLA value: {value!r}")
    locus, group, protein = match.groups()
    return f"HLA-{locus}*{group}:{protein}"


def load_training_rows(
    archive: Path, member: str, sheet: str
) -> tuple[list[dict[str, object]], str]:
    with ZipFile(archive) as handle:
        workbook_bytes = handle.read(member)
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    worksheet = workbook[sheet]
    iterator = worksheet.iter_rows(values_only=True)
    next(iterator)  # title row
    headers = [str(value) for value in next(iterator)]
    required = {"Mutant", "Allele", "Immunogenicity", "Random"}
    if required - set(headers):
        raise ValueError(f"{sheet} lacks columns: {sorted(required - set(headers))}")
    rows = []
    for values in iterator:
        row = dict(zip(headers, values, strict=True))
        if row["Mutant"] is None:
            continue
        label = int(row["Immunogenicity"])
        if label not in {0, 1}:
            raise ValueError(f"non-binary PRIME2 label: {row['Immunogenicity']!r}")
        rows.append(
            {
                "peptide": str(row["Mutant"]).strip().upper(),
                "hla": normalize_hla(row["Allele"]),
                "label": label,
                "random": int(row["Random"]),
            }
        )
    return rows, sha256(workbook_bytes)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--benchmark", type=Path, required=True)
    parser.add_argument("--supplement-zip", type=Path, required=True)
    parser.add_argument("--member", default="mmc6.xlsx")
    parser.add_argument("--sheet", default="TableS4")
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--summary", type=Path, required=True)
    parser.add_argument("--source-url", required=True)
    args = parser.parse_args()

    with args.benchmark.open(newline="", encoding="utf-8") as handle:
        benchmark = list(csv.DictReader(handle))
    training, workbook_hash = load_training_rows(args.supplement_zip, args.member, args.sheet)
    exact: dict[tuple[str, str], set[tuple[int, int]]] = defaultdict(set)
    peptide_only: dict[str, set[int]] = defaultdict(set)
    near_patterns: dict[tuple[str, int, str], set[str]] = defaultdict(set)
    for row in training:
        peptide = str(row["peptide"])
        hla = str(row["hla"])
        exact[(peptide, hla)].add((int(row["label"]), int(row["random"])))
        peptide_only[peptide].add(int(row["label"]))
        for index in range(len(peptide)):
            near_patterns[(hla, len(peptide), peptide[:index] + "*" + peptide[index + 1 :])].add(
                peptide
            )

    output_rows = []
    for row in benchmark:
        key = (row["peptide"].strip().upper(), row["hla"])
        exact_matches = exact.get(key, set())
        exact_labels = {label for label, _ in exact_matches}
        peptide_labels = peptide_only.get(key[0], set())
        near_candidates = set()
        for index in range(len(key[0])):
            pattern = key[0][:index] + "*" + key[0][index + 1 :]
            near_candidates.update(near_patterns.get((key[1], len(key[0]), pattern), set()))
        near_candidates.discard(key[0])
        benchmark_label = int(row["immunogenicity"])
        if exact_labels:
            classification = (
                "exact_label_conflict"
                if any(label != benchmark_label for label in exact_labels)
                else "exact_label_concordant"
            )
        elif peptide_labels:
            classification = "peptide_only_different_hla"
        else:
            classification = "no_overlap"
        output_rows.append(
            {
                "record_id": row["record_id"],
                "benchmark_label": benchmark_label,
                "exact_peptide_hla_in_prime2_train": int(bool(exact_labels)),
                "peptide_in_prime2_train": int(bool(peptide_labels)),
                "matched_training_labels": ";".join(map(str, sorted(exact_labels))),
                "prime2_random_flag": ";".join(
                    map(str, sorted({random_flag for _, random_flag in exact_matches}))
                ),
                "exact_bigmhc_im_trainval_overlap": int(
                    len(key[0]) in {9, 10}
                    and any(random_flag == 0 for _, random_flag in exact_matches)
                ),
                "near_hamming1_same_hla_prime2_train": int(bool(near_candidates)),
                "near_hamming1_training_peptide_count": len(near_candidates),
                "classification": classification,
            }
        )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(output_rows[0]), lineterminator="\n")
        writer.writeheader()
        writer.writerows(output_rows)

    classifications: dict[str, int] = defaultdict(int)
    for row in output_rows:
        classifications[str(row["classification"])] += 1
    summary = {
        "source_url": args.source_url,
        "supplement_archive_sha256": sha256(args.supplement_zip.read_bytes()),
        "workbook_member": args.member,
        "workbook_member_sha256": workbook_hash,
        "sheet": args.sheet,
        "prime2_training_rows": len(training),
        "prime2_unique_peptide_hla": len(exact),
        "prime2_unique_peptides": len(peptide_only),
        "benchmark_rows": len(benchmark),
        "benchmark_unique_peptide_hla": len({(r["peptide"], r["hla"]) for r in benchmark}),
        "benchmark_unique_peptides": len({r["peptide"] for r in benchmark}),
        "benchmark_record_classifications": dict(sorted(classifications.items())),
        "benchmark_records_exact_bigmhc_im_trainval_overlap": sum(
            int(row["exact_bigmhc_im_trainval_overlap"]) for row in output_rows
        ),
        "benchmark_records_near_hamming1_same_hla": sum(
            int(row["near_hamming1_same_hla_prime2_train"]) for row in output_rows
        ),
        "overlap_dimensions": {
            "exact_peptide": "checked",
            "exact_peptide_hla": "checked",
            "near_hamming1_same_hla_same_length": "checked",
            "mutation_identity": "unavailable_in_prime2_training_table",
            "patient_identity": "unavailable_in_prime2_training_table",
            "study_identity": "unavailable_in_prime2_training_table",
        },
        "interpretation": (
            "Exact peptide-HLA matches are direct training overlap for PRIME2 and must be "
            "excluded or reported separately. BigMHC im_trainval uses non-random PRIME1/2 "
            "peptides of length 9-10, so its exact overlaps are also training leakage."
        ),
    }
    args.summary.parent.mkdir(parents=True, exist_ok=True)
    args.summary.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(summary["benchmark_record_classifications"], sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
