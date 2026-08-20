#!/usr/bin/env python3
"""Build the Zhao 2026 administered-peptide vaccine benchmark from its workbook."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median

from openpyxl import load_workbook

MEMBER_SHA256 = "1fa76cf45435c39dc28e9d52e584d56938cee531dda953ad11cfbcc9c2617aea"
MEMBER = "Table1.xlsx"
SHEET = "S1"
DOI = "10.3389/fimmu.2026.1829509"
PMID = "42344930"
SOURCE_URL = "https://pmc.ncbi.nlm.nih.gov/articles/PMC13286890/"
SOURCE_DOWNLOAD = (
    "https://www.ebi.ac.uk/europepmc/webservices/rest/PMC13286890/supplementaryFiles"
)
SOURCE_ROWS = 2317
SOURCE_PATIENTS = 352
SOURCE_POSITIVES = 313
AMINO_ACIDS = set("ACDEFGHIKLMNPQRSTVWY")

CORE_FIELDS = [
    "record_id",
    "sample_id",
    "patient_id",
    "study_id",
    "cancer_type",
    "gene",
    "mutation",
    "wildtype_sequence",
    "mutant_sequence",
    "peptide",
    "peptide_length",
    "hla",
    "mhc_class",
    "assay_type",
    "immunogenicity",
    "presentation_evidence",
    "mass_spec_evidence",
    "clinical_context",
    "source_doi",
    "source_pmid",
    "source_url",
    "original_dataset",
    "evidence_level",
    "source_license",
    "source_checksum",
    "accessed_at",
]
EXTRA_FIELDS = [
    "source_row",
    "source_peptide_id",
    "elispot_ratio_source",
    "elispot_ratio_lower_bound",
]


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def scalar(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_hla(value: object) -> str:
    compact = re.sub(r"[^A-Z0-9]", "", scalar(value).upper().removeprefix("HLA"))
    match = re.fullmatch(r"([ABC])(\d{2})(\d{2})", compact)
    if not match:
        raise ValueError(f"unsupported class-I HLA value: {value!r}")
    locus, group, protein = match.groups()
    return f"HLA-{locus}*{group}:{protein}"


def stable_record_id(peptide_id: str, patient: str, peptide: str, hla: str) -> str:
    identity = f"{peptide_id}|{patient}|{peptide}|{hla}".encode()
    return f"zhao-vaccine-{hashlib.sha256(identity).hexdigest()[:16]}"


def build(workbook_path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook_bytes = workbook_path.read_bytes()
    observed_member = sha256(workbook_bytes)
    if observed_member != MEMBER_SHA256:
        raise ValueError(f"workbook checksum mismatch: {observed_member}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[SHEET]
    iterator = worksheet.iter_rows(values_only=True)
    title = next(iterator)[0]
    headers = [scalar(value) for value in next(iterator)]
    required = {
        "Peptide ID",
        "Patient ID",
        "Cancer type",
        "ELSPOT ratio",
        "mutation type",
        "peptide(mut)",
        "peptide(wt)",
        "Length",
        "HLA type",
    }
    if required - set(headers):
        raise ValueError(f"workbook lacks columns: {sorted(required - set(headers))}")

    current_patient = ""
    current_cancer = ""
    rows: list[dict[str, object]] = []
    exact_seen: set[tuple[str, str, str, float]] = set()
    duplicate_count = 0
    right_censored_ratio_count = 0
    for source_row, values in enumerate(iterator, start=3):
        source = dict(zip(headers, values, strict=True))
        if source["Patient ID"] is not None:
            current_patient = scalar(source["Patient ID"])
        if source["Cancer type"] is not None:
            current_cancer = scalar(source["Cancer type"])
        if not current_patient or not current_cancer:
            raise ValueError(f"row {source_row}: patient block lacks patient or cancer type")

        peptide_id = scalar(source["Peptide ID"])
        peptide = scalar(source["peptide(mut)"]).upper()
        wildtype = scalar(source["peptide(wt)"]).upper()
        hla = normalize_hla(source["HLA type"])
        ratio_source = scalar(source["ELSPOT ratio"])
        right_censored = ratio_source.startswith(("≥", ">="))
        ratio = float(ratio_source.removeprefix("≥").removeprefix(">="))
        right_censored_ratio_count += int(right_censored)
        declared_length = int(source["Length"])
        if not math.isfinite(ratio):
            raise ValueError(f"row {source_row}: nonfinite ELSPOT ratio")
        if len(peptide) != declared_length or not 8 <= len(peptide) <= 11:
            raise ValueError(f"row {source_row}: inconsistent peptide length")
        if set(peptide) - AMINO_ACIDS or set(wildtype) - AMINO_ACIDS:
            raise ValueError(f"row {source_row}: noncanonical amino acid")
        if scalar(source["mutation type"]).upper() != "SNV":
            raise ValueError(f"row {source_row}: non-SNV row in source S1")

        exact_key = (current_patient, peptide, hla, ratio)
        if exact_key in exact_seen:
            duplicate_count += 1
            continue
        exact_seen.add(exact_key)
        label = int(ratio >= 2.0)
        rows.append(
            {
                "record_id": stable_record_id(peptide_id, current_patient, peptide, hla),
                "sample_id": f"ZHAO-{current_patient}",
                "patient_id": f"ZHAO-{current_patient}",
                "study_id": "ZHAO_DC_VACCINE_2026",
                "cancer_type": current_cancer,
                "gene": "",
                "mutation": "SNV",
                "wildtype_sequence": wildtype,
                "mutant_sequence": peptide,
                "peptide": peptide,
                "peptide_length": len(peptide),
                "hla": hla,
                "mhc_class": "I",
                "assay_type": "post_vaccination_IFNG_ELISPOT",
                "immunogenicity": label,
                "presentation_evidence": "not_measured",
                "mass_spec_evidence": "not_measured",
                "clinical_context": "personalized_peptide_pulsed_DC_vaccine",
                "source_doi": DOI,
                "source_pmid": PMID,
                "source_url": SOURCE_URL,
                "original_dataset": "Zhao2026_DC_vaccine_short_SNV_peptides",
                "evidence_level": "administered_peptide_post_vaccination_T_cell_response",
                "source_license": "CC-BY-4.0",
                "source_checksum": MEMBER_SHA256,
                "accessed_at": "2026-08-20",
                "source_row": source_row,
                "source_peptide_id": peptide_id,
                "elispot_ratio_source": ratio_source,
                "elispot_ratio_lower_bound": f"{ratio:.12g}",
            }
        )

    if len(rows) != SOURCE_ROWS or duplicate_count:
        raise ValueError(f"unexpected row/duplicate count: {len(rows)} rows, {duplicate_count} dupes")
    if len({str(row["patient_id"]) for row in rows}) != SOURCE_PATIENTS:
        raise ValueError("unexpected patient count")
    if sum(int(row["immunogenicity"]) for row in rows) != SOURCE_POSITIVES:
        raise ValueError("unexpected positive count")
    if len({str(row["record_id"]) for row in rows}) != len(rows):
        raise ValueError("record identifiers are not unique")

    by_patient: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        by_patient[str(row["patient_id"])].append(row)
    summary = {
        "title": title,
        "source_url": SOURCE_DOWNLOAD,
        "article_doi": DOI,
        "workbook_member": MEMBER,
        "workbook_member_sha256": observed_member,
        "source_rows": len(rows),
        "output_rows": len(rows),
        "exact_duplicates_removed": duplicate_count,
        "right_censored_elispot_ratios": right_censored_ratio_count,
        "patients": len(by_patient),
        "positive_bearing_patients": sum(
            any(int(row["immunogenicity"]) for row in patient_rows)
            for patient_rows in by_patient.values()
        ),
        "positives": sum(int(row["immunogenicity"]) for row in rows),
        "negatives": sum(not int(row["immunogenicity"]) for row in rows),
        "peptide_lengths": dict(sorted(Counter(row["peptide_length"] for row in rows).items())),
        "hla_alleles": len({str(row["hla"]) for row in rows}),
        "cancer_types": len({str(row["cancer_type"]) for row in rows}),
        "candidates_per_patient": {
            "min": min(map(len, by_patient.values())),
            "median": median(map(len, by_patient.values())),
            "max": max(map(len, by_patient.values())),
        },
        "label_rule": "ELSPOT ratio >= 2.0",
        "endpoint": "post-vaccination IFN-gamma ELISPOT response to administered peptide",
    }
    return rows, summary


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--summary", type=Path, required=True)
    args = parser.parse_args()
    rows, summary = build(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=CORE_FIELDS + EXTRA_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    args.summary.parent.mkdir(parents=True, exist_ok=True)
    args.summary.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(summary, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
